import os
import openpyxl as opx

# create Excel workbook and sheet with index 0 
wb = opx.Workbook()
imp_sh = wb.create_sheet('compile', 0)

j = 1
imp_sh['A' + str(j)] = 'Customer'
imp_sh['B' + str(j)] = 'Contract number'
imp_sh['C' + str(j)] = 'Date contract'

#loop by folder with raw files 
for files in os.listdir('.' + '\\raw'):
    sheet = opx.load_workbook('.' + '\\raw\\' + files).worksheets[0]
    i = 0
    for row in sheet.rows:
        i += 1
        if i > 3: # condition for start to insert 
            j += 1
            imp_sh['A' + str(j)] = row[1].value
            imp_sh['B' + str(j)] = row[2].value
            imp_sh['C' + str(j)] = row[3].value
            imp_sh['D' + str(j)] = files
wb.save('compile.xlsx')
